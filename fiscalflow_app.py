import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import time
from datetime import datetime

# ─────────────────────────────────────────
#  CONFIG DA PÁGINA
# ─────────────────────────────────────────
st.set_page_config(
    page_title="FiscalFlow — Processador NF-e",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────
#  ESTILOS CUSTOMIZADOS
# ─────────────────────────────────────────
st.markdown("""
<style>
    /* Fundo e fonte */
    .stApp { background-color: #0e0f13; }
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; color: #e8eaf0; }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background-color: #16181f;
        border-right: 1px solid #2a2d38;
    }

    /* Cards de métricas */
    div[data-testid="metric-container"] {
        background: #16181f;
        border: 1px solid #2a2d38;
        border-radius: 12px;
        padding: 16px;
    }
    div[data-testid="stMetricLabel"] p,
    div[data-testid="stMetricLabel"] {
        color: #c0c4d0 !important;
        font-size: 13px !important;
        font-weight: 600 !important;
        opacity: 1 !important;
        visibility: visible !important;
    }
    div[data-testid="stMetricValue"] div,
    div[data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-size: 26px !important;
        font-weight: 700 !important;
    }

    /* Títulos gerais e textos */
    h1, h2, h3, h4 { color: #ffffff !important; }
    p { color: #c0c4d0 !important; }

    /* Caption / subtítulos */
    div[data-testid="stCaptionContainer"] p,
    .stCaption { color: #8a8fa0 !important; }

    /* Markdown textos */
    div[data-testid="stMarkdownContainer"] p { color: #c0c4d0 !important; }
    div[data-testid="stMarkdownContainer"] h1,
    div[data-testid="stMarkdownContainer"] h2,
    div[data-testid="stMarkdownContainer"] h3 { color: #ffffff !important; }
    div[data-testid="stMarkdownContainer"] strong { color: #ffffff !important; }

    /* Labels de inputs, file uploader, selectbox, number input */
    label, .stSelectbox label, .stNumberInput label,
    div[data-testid="stFileUploaderDropzoneInstructions"],
    div[data-testid="stFileUploader"] label {
        color: #c0c4d0 !important;
        font-size: 14px !important;
        font-weight: 500 !important;
    }

    /* Expander títulos */
    details summary p,
    div[data-testid="stExpander"] summary p {
        color: #ffffff !important;
        font-weight: 600 !important;
    }

    /* Info / warning / success boxes */
    div[data-testid="stAlert"] p { color: #ffffff !important; }

    /* Sidebar textos */
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] span {
        color: #c0c4d0 !important;
    }
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {
        color: #ffffff !important;
    }

    /* Radio buttons */
    div[data-testid="stRadio"] label p { color: #c0c4d0 !important; }
    div[data-testid="stRadio"] label:hover p { color: #ffffff !important; }

    /* Tabela / dataframe */
    div[data-testid="stDataFrame"] { color: #e8eaf0 !important; }

    /* Lista de arquivos do uploader */
    div[data-testid="stFileUploaderFileData"] span,
    div[data-testid="stFileUploaderFileData"] p,
    div[data-testid="stFileUploaderFileData"] small,
    div[data-testid="stFileUploader"] span,
    div[data-testid="uploadedFileData"] span,
    div[data-testid="uploadedFileName"],
    div[data-testid="stFileUploaderDeleteBtn"],
    section[data-testid="stFileUploaderDropzone"] span,
    div[class*="uploadedFile"] span,
    div[class*="fileUploader"] span {
        color: #c0c4d0 !important;
        opacity: 1 !important;
    }
    /* Texto "Exibindo página X de Y" */
    div[data-testid="stFileUploader"] p,
    div[data-testid="stFileUploader"] small {
        color: #8a8fa0 !important;
        opacity: 1 !important;
    }

    /* Botão primário */
    .stButton > button {
        background: #00e5a0;
        color: #000000 !important;
        font-weight: 700 !important;
        font-size: 15px !important;
        border: none;
        border-radius: 8px;
        padding: 10px 20px;
        width: 100%;
        transition: all 0.2s;
    }
    .stButton > button p,
    .stButton > button span {
        color: #000000 !important;
        font-weight: 700 !important;
    }
    .stButton > button:hover {
        background: #00ffb3;
        color: #000000 !important;
        transform: translateY(-1px);
        box-shadow: 0 4px 20px rgba(0,229,160,0.3);
    }

    /* Cards de upload da sidebar (NCM ST e MVA) */
    section[data-testid="stSidebar"] div[data-testid="stFileUploaderDropzone"] {
        background: #1e2029 !important;
        border: 1px solid #2a2d38 !important;
        border-radius: 8px !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stFileUploaderDropzone"] span,
    section[data-testid="stSidebar"] div[data-testid="stFileUploaderDropzone"] p,
    section[data-testid="stSidebar"] div[data-testid="stFileUploaderDropzoneInstructions"] span,
    section[data-testid="stSidebar"] div[data-testid="stFileUploaderDropzoneInstructions"] p {
        color: #000000 !important;
        font-weight: 600 !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stFileUploader"] button {
        color: #c0c4d0 !important;
        border-color: #2a2d38 !important;
        background: #16181f !important;
    }
    /* Área de drop do uploader principal (central) */
    div[data-testid="stFileUploaderDropzone"] {
        background: #1e2029 !important;
        border-radius: 10px !important;
        border: 1px solid #2a2d38 !important;
    }
    div[data-testid="stFileUploaderDropzoneInstructions"] span,
    div[data-testid="stFileUploaderDropzoneInstructions"] p,
    div[data-testid="stFileUploaderDropzoneInstructions"] small,
    div[data-testid="stFileUploaderDropzone"] span,
    div[data-testid="stFileUploaderDropzone"] p,
    div[data-testid="stFileUploaderDropzone"] small,
    div[data-testid="stFileUploaderDropzone"] button {
        color: #000000 !important;
        font-weight: 500 !important;
        background: transparent !important;
    }
    /* Botão "Procurar arquivos" */
    div[data-testid="stFileUploaderDropzone"] button {
        background: #2a2d38 !important;
        color: #ffffff !important;
        border: 1px solid #3a3d48 !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
    }
    div[data-testid="stFileUploaderDropzone"] button:hover {
        background: #3a3d48 !important;
        color: #ffffff !important;
    }

    /* Upload zone */
    div[data-testid="stFileUploader"] {
        background: #16181f;
        border: 2px dashed #2a2d38;
        border-radius: 16px;
        padding: 8px;
    }

    /* Inputs / selects */
    .stNumberInput input, .stSelectbox select {
        background: #1e2029 !important;
        border: 1px solid #2a2d38 !important;
        color: #e8eaf0 !important;
        border-radius: 8px !important;
    }

    /* Títulos */
    h1 { font-size: 26px !important; font-weight: 700 !important; letter-spacing: -0.5px; }
    h2 { font-size: 18px !important; font-weight: 700 !important; }
    h3 { font-size: 15px !important; font-weight: 600 !important; color: #6b7080 !important; }

    /* Tabela */
    .stDataFrame { border-radius: 12px; overflow: hidden; }

    /* Alerta de sucesso */
    .success-box {
        background: rgba(0,229,160,0.08);
        border: 1px solid rgba(0,229,160,0.25);
        border-radius: 12px;
        padding: 20px 24px;
        margin: 8px 0;
    }

    /* Tag chips */
    .chip {
        display: inline-block;
        background: rgba(0,229,160,0.1);
        color: #00e5a0;
        border: 1px solid rgba(0,229,160,0.2);
        padding: 3px 10px;
        border-radius: 20px;
        font-size: 12px;
        font-family: monospace;
    }
    .chip-warn {
        background: rgba(255,107,53,0.1);
        color: #ff6b35;
        border-color: rgba(255,107,53,0.2);
    }
    .chip-blue {
        background: rgba(77,159,255,0.1);
        color: #4d9fff;
        border-color: rgba(77,159,255,0.2);
    }

    /* Botão secundário (Limpar seleção, Baixar Excel) */
    button[kind="secondary"],
    div[data-testid="stDownloadButton"] button,
    .stButton > button[kind="secondary"] {
        background: #000000 !important;
        color: #ffffff !important;
        border: 1px solid #3a3d48 !important;
        font-weight: 600 !important;
    }
    div[data-testid="stDownloadButton"] button:hover,
    .stButton > button[kind="secondary"]:hover {
        background: #1e2029 !important;
        color: #ffffff !important;
        border-color: #00e5a0 !important;
    }
    /* Botão Procurar arquivos */
    div[data-testid="stFileUploaderDropzone"] button {
        background: #000000 !important;
        color: #ffffff !important;
        border: 1px solid #3a3d48 !important;
        font-weight: 600 !important;
    }
    div[data-testid="stFileUploaderDropzone"] button:hover {
        background: #1e2029 !important;
        border-color: #00e5a0 !important;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
#  PARSER NF-e
# ─────────────────────────────────────────
NS = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

def _get(el, path):
    found = el.find(path, NS)
    return found.text if found is not None else ''

def parse_nfe(xml_bytes):
    root = ET.fromstring(xml_bytes)
    nfe = root.find('.//nfe:infNFe', NS)
    if nfe is None:
        return []

    ide     = nfe.find('nfe:ide', NS)
    emit    = nfe.find('nfe:emit', NS)
    dest    = nfe.find('nfe:dest', NS)
    total   = nfe.find('nfe:total/nfe:ICMSTot', NS)
    transp  = nfe.find('nfe:transp', NS)
    infAdic = nfe.find('nfe:infAdic', NS)
    protNFe = root.find('.//nfe:protNFe/nfe:infProt', NS)

    nfe_id = nfe.get('Id', '')
    chave  = nfe_id.replace('NFe', '') if nfe_id else ''

    header = {
        'chave_nfe':    chave,
        'num_nf':       _get(ide, 'nfe:nNF'),
        'serie':        _get(ide, 'nfe:serie'),
        'data_emissao': _get(ide, 'nfe:dhEmi'),
        'nat_op':       _get(ide, 'nfe:natOp'),
        'emit_cnpj':    _get(emit, 'nfe:CNPJ'),
        'emit_nome':    _get(emit, 'nfe:xNome'),
        'emit_uf':      _get(emit, 'nfe:enderEmit/nfe:UF'),
        'dest_cnpj':    _get(dest, 'nfe:CNPJ') or _get(dest, 'nfe:CPF'),
        'dest_nome':    _get(dest, 'nfe:xNome'),
        'dest_uf':      _get(dest, 'nfe:enderDest/nfe:UF'),
        'dest_mun':     _get(dest, 'nfe:enderDest/nfe:xMun'),
        'vBC_total':    _get(total, 'nfe:vBC')    if total is not None else '',
        'vICMS_total':  _get(total, 'nfe:vICMS')  if total is not None else '',
        'vIPI_total':   _get(total, 'nfe:vIPI')   if total is not None else '',
        'vPIS_total':   _get(total, 'nfe:vPIS')   if total is not None else '',
        'vCOFINS_total':_get(total, 'nfe:vCOFINS')if total is not None else '',
        'vProd_total':  _get(total, 'nfe:vProd')  if total is not None else '',
        'vNF_total':    _get(total, 'nfe:vNF')    if total is not None else '',
        'vFrete':       _get(total, 'nfe:vFrete') if total is not None else '',
        'vDesc':        _get(total, 'nfe:vDesc')  if total is not None else '',
        'transp_nome':  _get(transp, 'nfe:transporta/nfe:xNome') if transp is not None else '',
        'transp_uf':    _get(transp, 'nfe:transporta/nfe:UF')    if transp is not None else '',
        'inf_compl':    _get(infAdic, 'nfe:infCpl') if infAdic is not None else '',
        'nProt':        _get(protNFe, 'nfe:nProt')   if protNFe is not None else '',
        'dhRecbto':     _get(protNFe, 'nfe:dhRecbto')if protNFe is not None else '',
        'cStat':        _get(protNFe, 'nfe:cStat')   if protNFe is not None else '',
        'xMotivo':      _get(protNFe, 'nfe:xMotivo') if protNFe is not None else '',
    }

    rows = []
    for det in nfe.findall('nfe:det', NS):
        prod = det.find('nfe:prod', NS)
        imp  = det.find('nfe:imposto', NS)

        # ICMS
        icms_el = None
        icms_cst = icms_orig = icms_vbc = icms_pICMS = icms_vICMS = ''
        if imp is not None:
            icms_g = imp.find('nfe:ICMS', NS)
            if icms_g is not None:
                for child in icms_g: icms_el = child; break
            if icms_el is not None:
                icms_cst   = icms_el.findtext('nfe:CST', '', NS) or icms_el.findtext('nfe:CSOSN', '', NS)
                icms_orig  = _get(icms_el, 'nfe:orig')
                icms_vbc   = _get(icms_el, 'nfe:vBC')
                icms_pICMS = _get(icms_el, 'nfe:pICMS')
                icms_vICMS = _get(icms_el, 'nfe:vICMS')

        # IPI
        ipi_cst = ipi_vbc = ipi_pIPI = ipi_vIPI = ''
        if imp is not None:
            ipi_g = imp.find('nfe:IPI', NS)
            if ipi_g is not None:
                for el in ipi_g:
                    if el.tag.endswith('IPITrib') or el.tag.endswith('IPINT'):
                        ipi_cst = _get(el, 'nfe:CST'); ipi_vbc = _get(el, 'nfe:vBC')
                        ipi_pIPI = _get(el, 'nfe:pIPI'); ipi_vIPI = _get(el, 'nfe:vIPI')

        # PIS
        pis_cst = pis_vbc = pis_pPIS = pis_vPIS = ''
        if imp is not None:
            pis_g = imp.find('nfe:PIS', NS)
            if pis_g is not None:
                for el in pis_g:
                    pis_cst = _get(el, 'nfe:CST'); pis_vbc = _get(el, 'nfe:vBC')
                    pis_pPIS = _get(el, 'nfe:pPIS'); pis_vPIS = _get(el, 'nfe:vPIS')

        # COFINS
        cof_cst = cof_vbc = cof_pCOFINS = cof_vCOFINS = ''
        if imp is not None:
            cof_g = imp.find('nfe:COFINS', NS)
            if cof_g is not None:
                for el in cof_g:
                    cof_cst = _get(el, 'nfe:CST'); cof_vbc = _get(el, 'nfe:vBC')
                    cof_pCOFINS = _get(el, 'nfe:pCOFINS'); cof_vCOFINS = _get(el, 'nfe:vCOFINS')

        row = {**header}
        row.update({
            'n_item':    det.get('nItem', ''),
            'cod_prod':  _get(prod, 'nfe:cProd'),
            'desc_prod': _get(prod, 'nfe:xProd'),
            'ncm':       _get(prod, 'nfe:NCM'),
            'cest':      _get(prod, 'nfe:CEST'),
            'cfop':      _get(prod, 'nfe:CFOP'),
            'un_com':    _get(prod, 'nfe:uCom'),
            'qtd':       _get(prod, 'nfe:qCom'),
            'vUnCom':    _get(prod, 'nfe:vUnCom'),
            'vProd':     _get(prod, 'nfe:vProd'),
            'xPed':      _get(prod, 'nfe:xPed'),
            'nItemPed':  _get(prod, 'nfe:nItemPed'),
            'icms_orig': icms_orig, 'icms_cst': icms_cst,
            'icms_vBC':  icms_vbc,  'icms_pICMS': icms_pICMS, 'icms_vICMS': icms_vICMS,
            'ipi_cst':   ipi_cst,   'ipi_vBC': ipi_vbc, 'ipi_pIPI': ipi_pIPI, 'ipi_vIPI': ipi_vIPI,
            'pis_cst':   pis_cst,   'pis_vBC': pis_vbc, 'pis_pPIS': pis_pPIS, 'pis_vPIS': pis_vPIS,
            'cofins_cst':cof_cst,   'cofins_vBC': cof_vbc, 'cofins_pCOFINS': cof_pCOFINS, 'cofins_vCOFINS': cof_vCOFINS,
        })
        rows.append(row)
    return rows

# ─────────────────────────────────────────
#  CÁLCULO ICMS-ST
# ─────────────────────────────────────────
def calcular_st(df, ncm_st_set, mva_map, aliquota_interna):
    # Garante que todas as colunas numéricas existem, criando com 0 se ausentes
    for c in ['vProd', 'icms_pICMS', 'icms_vICMS', 'ipi_vIPI']:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    # Garante que coluna ncm existe
    if 'ncm' not in df.columns:
        df['ncm'] = ''

    df['ncm_str'] = df['ncm'].astype(str).str.strip()
    df['SITUAÇÃO ST'] = df['ncm_str'].apply(lambda x: 'COM ST' if x in ncm_st_set else 'NORMAL')

    df['icms_pICMS_int'] = df['icms_pICMS'].apply(lambda x: int(float(x)) if float(x) > 0 else 0)
    df['MVA'] = df['icms_pICMS_int'].map(mva_map)

    mask = df['SITUAÇÃO ST'] == 'COM ST'
    df['VALOR TOTAL']      = None
    df['BASE ICMS ST']     = None
    df['VALOR DO ICMS ST'] = None

    aliq = aliquota_interna / 100
    df.loc[mask, 'VALOR TOTAL']      = df.loc[mask, 'vProd'] + df.loc[mask, 'ipi_vIPI']
    df.loc[mask, 'BASE ICMS ST']     = (df.loc[mask, 'VALOR TOTAL'] * df.loc[mask, 'MVA']) + df.loc[mask, 'VALOR TOTAL']
    df.loc[mask, 'VALOR DO ICMS ST'] = (df.loc[mask, 'BASE ICMS ST'] * aliq) - df.loc[mask, 'icms_vICMS']

    return df

# ─────────────────────────────────────────
#  GERADOR DE EXCEL
# ─────────────────────────────────────────
def gerar_excel(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Itens NF-e"

    columns = [
        ('chave_nfe','Chave NF-e'), ('num_nf','Número NF'), ('serie','Série'),
        ('data_emissao','Data Emissão'), ('dhRecbto','Recebimento SEFAZ'),
        ('nat_op','Natureza Operação'), ('emit_cnpj','CNPJ Emitente'),
        ('emit_nome','Razão Social Emitente'), ('emit_uf','UF Emitente'),
        ('dest_cnpj','CNPJ Destinatário'), ('dest_nome','Razão Social Destinatário'),
        ('dest_uf','UF Dest'), ('dest_mun','Município Dest'),
        ('transp_nome','Transportadora'), ('vProd_total','Valor Produtos NF'),
        ('vIPI_total','IPI Total NF'), ('vICMS_total','ICMS Total NF'),
        ('vNF_total','Valor Total NF'), ('nProt','Protocolo'), ('cStat','Status'),
        ('n_item','Nº Item'), ('xPed','Pedido'), ('cod_prod','Código Produto'),
        ('desc_prod','Descrição Produto'), ('ncm','NCM'), ('cest','CEST'),
        ('cfop','CFOP'), ('un_com','Unidade'), ('qtd','Quantidade'),
        ('vUnCom','Valor Unitário'), ('vProd','Valor Total Item'),
        ('icms_orig','ICMS Origem'), ('icms_cst','ICMS CST'), ('icms_vBC','ICMS BC'),
        ('icms_pICMS','ICMS %'), ('icms_vICMS','ICMS Valor'),
        ('ipi_cst','IPI CST'), ('ipi_vBC','IPI BC'), ('ipi_pIPI','IPI %'), ('ipi_vIPI','IPI Valor'),
        ('pis_cst','PIS CST'), ('pis_vBC','PIS BC'), ('pis_pPIS','PIS %'), ('pis_vPIS','PIS Valor'),
        ('cofins_cst','COFINS CST'), ('cofins_vBC','COFINS BC'), ('cofins_pCOFINS','COFINS %'), ('cofins_vCOFINS','COFINS Valor'),
        ('SITUAÇÃO ST','SITUAÇÃO ST'), ('MVA','MVA'),
        ('VALOR TOTAL','VALOR TOTAL'), ('BASE ICMS ST','BASE ICMS ST'), ('VALOR DO ICMS ST','VALOR DO ICMS ST'),
    ]

    h_font    = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    h_fill    = PatternFill('solid', start_color='1F4E79')
    h_fill_st = PatternFill('solid', start_color='833C00')
    d_font    = Font(name='Arial', size=9)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    st_keys   = {'SITUAÇÃO ST','MVA','VALOR TOTAL','BASE ICMS ST','VALOR DO ICMS ST'}
    num_keys  = {'qtd','vUnCom','vProd','vProd_total','vIPI_total','vICMS_total','vNF_total',
                 'icms_vBC','icms_pICMS','icms_vICMS','ipi_vBC','ipi_pIPI','ipi_vIPI',
                 'pis_vBC','pis_pPIS','pis_vPIS','cofins_vBC','cofins_pCOFINS','cofins_vCOFINS',
                 'VALOR TOTAL','BASE ICMS ST','VALOR DO ICMS ST','MVA'}
    st_fill   = PatternFill('solid', start_color='FCE4D6')

    for ci, (key, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = h_font
        cell.fill = h_fill_st if key in st_keys else h_fill
        cell.alignment = center


    for ri, (_, row_data) in enumerate(df.iterrows(), 2):
        for ci, (key, _) in enumerate(columns, 1):
            val = row_data.get(key, None)
            try:
                if val is not None and not isinstance(val, str) and pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            if val is not None and key in num_keys:
                try: val = float(val)
                except: pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = d_font
            cell.alignment = Alignment(vertical='center')
            if isinstance(val, float):
                cell.number_format = '0.0000' if key == 'MVA' else '#,##0.00'
            if key in st_keys and val is not None:
                cell.fill = st_fill

    widths = {'chave_nfe':50,'num_nf':12,'serie':8,'data_emissao':22,'dhRecbto':22,
              'nat_op':30,'emit_cnpj':18,'emit_nome':35,'dest_cnpj':18,'dest_nome':35,
              'desc_prod':35,'SITUAÇÃO ST':14,'MVA':12,
              'VALOR TOTAL':16,'BASE ICMS ST':16,'VALOR DO ICMS ST':18}
    for ci, (key, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(key, 13)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 35

    # Aba resumo
    ws2 = wb.create_sheet("Resumo por NF")
    h2  = ['Número NF','Série','Data Emissão','Emitente','Destinatário','UF Dest',
           'Qtd Itens','Itens COM ST','Itens NORMAL','Valor Produtos',
           'IPI Total','ICMS Total','Valor Total NF','ICMS-ST Total']
    for ci, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = h_font; cell.fill = h_fill; cell.alignment = center

    from collections import OrderedDict
    groups = OrderedDict()
    for _, r in df.iterrows():
        k = r['num_nf']
        if k not in groups: groups[k] = []
        groups[k].append(r)

    for ri2, (_, items) in enumerate(groups.items(), 2):
        f   = items[0]
        fv  = lambda k: float(f.get(k, 0) or 0)
        n_st= sum(1 for i in items if i.get('SITUAÇÃO ST') == 'COM ST')
        st_total = sum(float(i.get('VALOR DO ICMS ST') or 0) for i in items if i.get('SITUAÇÃO ST') == 'COM ST')
        vals = [f['num_nf'], f['serie'], f['data_emissao'], f['emit_nome'], f['dest_nome'], f['dest_uf'],
                len(items), n_st, len(items)-n_st,
                fv('vProd_total'), fv('vIPI_total'), fv('vICMS_total'), fv('vNF_total'), st_total]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(row=ri2, column=ci, value=val)
            cell.font = Font(name='Arial', size=9)
            if isinstance(val, float): cell.number_format = '#,##0.00'

    for ci in range(1, len(h2)+1):
        ws2.column_dimensions[get_column_letter(ci)].width = 20
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = ws2.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────────
if 'aliquota_st' not in st.session_state:
    st.session_state.aliquota_st = 20.0

# ─────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚡ FiscalFlow")
    st.markdown("---")

    pagina = st.radio(
        "Navegação",
        ["📤 Processar XMLs", "⚙️ Configurações"],
        label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown("### Tabelas ativas")

    ncm_file = st.file_uploader("NCM ST (.xlsx)", type=['xlsx'], key="ncm_upload")
    mva_file = st.file_uploader("MVA (.xlsx)", type=['xlsx'], key="mva_upload")

    if ncm_file:
        st.success(f"✅ NCM ST carregada")
    else:
        st.warning("⚠️ Suba a tabela NCM ST")

    if mva_file:
        st.success(f"✅ MVA carregada")
    else:
        st.warning("⚠️ Suba a tabela MVA")

    st.markdown("---")
    st.markdown("### Alíquota ICMS-ST")
    aliquota = st.number_input(
        "Alíquota interna (%)",
        min_value=1.0, max_value=30.0,
        value=st.session_state.aliquota_st,
        step=0.5, format="%.1f"
    )
    st.session_state.aliquota_st = aliquota
    st.caption(f"Fórmula: (Base × {aliquota:.1f}%) − ICMS Valor")

# ─────────────────────────────────────────
#  PÁGINA: PROCESSAR XMLs
# ─────────────────────────────────────────
if pagina == "📤 Processar XMLs":
    st.title("Processar NF-e")
    st.caption("Faça upload dos XMLs para extrair dados e calcular ICMS-ST automaticamente.")

    st.markdown("---")

    # Upload XMLs
    # Botão limpar
    col_clear, col_space = st.columns([1, 4])
    with col_clear:
        if st.button("🗑️ Limpar seleção", help="Remove todos os XMLs selecionados para subir um novo lote"):
            st.session_state['upload_key'] = st.session_state.get('upload_key', 0) + 1
            st.rerun()

    uploaded_xmls = st.file_uploader(
        "Arraste os XMLs das NF-e aqui",
        type=['xml'],
        accept_multiple_files=True,
        help="Suporte a NF-e modelo 55 versão 4.00. Múltiplos arquivos de uma vez.",
        key=f"xml_uploader_{st.session_state.get('upload_key', 0)}"
    )

    if uploaded_xmls:
        st.markdown(f"**{len(uploaded_xmls)} arquivo(s) selecionado(s)**")

        col_btn1, col_btn2 = st.columns([1, 3])
        with col_btn1:
            processar = st.button("▶ Processar agora", type="primary")

        if processar:
            # Valida tabelas
            if not ncm_file or not mva_file:
                st.error("⚠️ Suba as tabelas NCM ST e MVA na barra lateral antes de processar.")
                st.stop()

            # Carrega tabelas de referência
            ncm_st_df  = pd.read_excel(ncm_file)
            mva_df     = pd.read_excel(mva_file)
            ncm_st_set = set(ncm_st_df['NCM'].astype(str).str.strip().tolist())
            mva_map    = dict(zip(mva_df['Alíquota Interestadual'].astype(int), mva_df['MVA']))

            # Processa
            progress = st.progress(0, text="Iniciando...")
            all_rows = []
            erros    = []

            for i, f in enumerate(uploaded_xmls):
                progress.progress((i + 1) / len(uploaded_xmls),
                                  text=f"Lendo {f.name} ({i+1}/{len(uploaded_xmls)})")
                try:
                    rows = parse_nfe(f.read())
                    all_rows.extend(rows)
                except Exception as e:
                    erros.append(f"{f.name}: {str(e)}")

            progress.progress(1.0, text="Calculando ICMS-ST...")
            time.sleep(0.3)

            df = pd.DataFrame(all_rows)
            df = calcular_st(df, ncm_st_set, mva_map, st.session_state.aliquota_st)

            progress.empty()

            # Erros de leitura
            if erros:
                st.warning(f"⚠️ {len(erros)} arquivo(s) com erro:")
                for e in erros:
                    st.caption(f"• {e}")

            # Resultados
            mask       = df['SITUAÇÃO ST'] == 'COM ST'
            num_nfs    = df['num_nf'].nunique()
            num_itens  = len(df)
            num_st     = mask.sum()
            icms_st    = df.loc[mask, 'VALOR DO ICMS ST'].sum()

            st.markdown("### ✅ Processamento concluído")
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Notas Fiscais",  num_nfs)
            m2.metric("Itens totais",   num_itens)
            m3.metric("COM ST",         num_st)
            m4.metric("NORMAL",         num_itens - num_st)
            m5.metric("ICMS-ST Total",  f"R$ {icms_st:,.2f}")

            # Gera Excel
            excel_buf = gerar_excel(df)
            nome_arquivo = f"NF-e_ST_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            st.download_button(
                label="⬇ Baixar Excel",
                data=excel_buf,
                file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )



            # Preview agrupado por NF — tabela estilizada
            st.markdown("### Prévia — Resumo por Nota Fiscal")
            preview_df = df[mask].copy()
            for col in ['VALOR TOTAL', 'BASE ICMS ST', 'VALOR DO ICMS ST']:
                preview_df[col] = pd.to_numeric(preview_df[col], errors='coerce').fillna(0)

            resumo = preview_df.groupby('num_nf', as_index=False).agg(
                VALOR_TOTAL   =('VALOR TOTAL',      'sum'),
                BASE_ICMS_ST  =('BASE ICMS ST',     'sum'),
                VALOR_ICMS_ST =('VALOR DO ICMS ST', 'sum'),
            ).rename(columns={
                'num_nf':       'Número NF',
                'VALOR_TOTAL':  'VALOR TOTAL',
                'BASE_ICMS_ST': 'BASE ICMS ST',
                'VALOR_ICMS_ST':'VALOR DO ICMS ST',
            })

            total_row = {
                'Número NF':       'Total Geral',
                'VALOR TOTAL':      resumo['VALOR TOTAL'].sum(),
                'BASE ICMS ST':     resumo['BASE ICMS ST'].sum(),
                'VALOR DO ICMS ST': resumo['VALOR DO ICMS ST'].sum(),
            }

            # Gera HTML da tabela
            def fmt(v):
                try: return "R$ {:,.2f}".format(float(v))
                except: return str(v)

            rows_html = ""
            for i, row in resumo.iterrows():
                bg = "#1e2029" if i % 2 == 0 else "#16181f"
                nf   = str(row["Número NF"])
                vt   = fmt(row["VALOR TOTAL"])
                bi   = fmt(row["BASE ICMS ST"])
                vi   = fmt(row["VALOR DO ICMS ST"])
                rows_html += (
                    "<tr style='background:" + bg + ";'>"
                    "<td style='padding:8px 14px;color:#e8eaf0;font-size:13px;'>" + nf + "</td>"
                    "<td style='padding:8px 14px;color:#e8eaf0;font-size:13px;text-align:right;'>" + vt + "</td>"
                    "<td style='padding:8px 14px;color:#e8eaf0;font-size:13px;text-align:right;'>" + bi + "</td>"
                    "<td style='padding:8px 14px;color:#00e5a0;font-size:13px;text-align:right;font-weight:600;'>" + vi + "</td>"
                    "</tr>"
                )

            tg_vt = fmt(total_row["VALOR TOTAL"])
            tg_bi = fmt(total_row["BASE ICMS ST"])
            tg_vi = fmt(total_row["VALOR DO ICMS ST"])

            table_html = (
                "<div style='border-radius:12px;overflow:hidden;border:1px solid #2a2d38;margin-top:8px;'>"
                "<table style='width:100%;border-collapse:collapse;font-family:Arial,sans-serif;'>"
                "<thead><tr style='background:#1F4E79;'>"
                "<th style='padding:10px 14px;color:#ffffff;font-size:12px;text-align:left;'>NÚMERO NF</th>"
                "<th style='padding:10px 14px;color:#ffffff;font-size:12px;text-align:right;'>VALOR TOTAL</th>"
                "<th style='padding:10px 14px;color:#ffffff;font-size:12px;text-align:right;'>BASE ICMS ST</th>"
                "<th style='padding:10px 14px;color:#ffffff;font-size:12px;text-align:right;'>VALOR DO ICMS ST</th>"
                "</tr></thead>"
                "<tbody>" + rows_html + "</tbody>"
                "<tfoot><tr style='background:#0e0f13;border-top:2px solid #2a2d38;'>"
                "<td style='padding:10px 14px;color:#ffffff;font-size:13px;font-weight:700;'>Total Geral</td>"
                "<td style='padding:10px 14px;color:#ffffff;font-size:13px;font-weight:700;text-align:right;'>" + tg_vt + "</td>"
                "<td style='padding:10px 14px;color:#ffffff;font-size:13px;font-weight:700;text-align:right;'>" + tg_bi + "</td>"
                "<td style='padding:10px 14px;color:#00e5a0;font-size:13px;font-weight:700;text-align:right;'>" + tg_vi + "</td>"
                "</tr></tfoot>"
                "</table></div>"
            )

            st.markdown(table_html, unsafe_allow_html=True)



# ─────────────────────────────────────────
#  PÁGINA: CONFIGURAÇÕES
# ─────────────────────────────────────────
elif pagina == "⚙️ Configurações":
    st.title("Configurações")

    st.markdown("### Fórmula de Cálculo ICMS-ST")
    aliq = st.session_state.aliquota_st
    st.code(f"""
VALOR TOTAL      = Valor Total Item + IPI Valor
BASE ICMS ST     = (VALOR TOTAL × MVA) + VALOR TOTAL
VALOR DO ICMS ST = (BASE ICMS ST × {aliq:.1f}%) − ICMS Valor
    """, language="text")

    st.markdown("### Alterar Alíquota Interna")
    nova_aliq = st.number_input(
        "Alíquota interna ICMS-ST (%)",
        min_value=1.0, max_value=30.0,
        value=st.session_state.aliquota_st,
        step=0.5, format="%.1f",
        key="config_aliq"
    )
    if st.button("💾 Salvar alíquota"):
        st.session_state.aliquota_st = nova_aliq
        st.success(f"✅ Alíquota atualizada para {nova_aliq:.1f}%")

    st.markdown("### Sobre as Tabelas")
    st.info("""
**NCM ST** — Lista de NCMs sujeitos a Substituição Tributária. 
Suba sempre pela barra lateral. Duplicatas são tratadas automaticamente.

**MVA** — Margem de Valor Agregado por alíquota interestadual (4%, 7%, 12%).
Suba sempre pela barra lateral.
    """)
